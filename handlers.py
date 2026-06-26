import io
import re
import logging
import tempfile
import os
from datetime import date, datetime, timedelta

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes, ConversationHandler

from database import Database
from keyboards import (
    main_reply_keyboard, main_menu_keyboard, analytics_keyboard,
    reports_keyboard, budget_menu_keyboard, settings_keyboard,
    help_keyboard, expense_confirmation_keyboard, reminder_keyboard,
    export_keyboard, back_button
)
from utils import (
    format_date, format_currency, get_category_emoji, auto_detect_category,
    parse_expense, fun_comparison, progress_bar, get_date_range, get_month_name
)
from achievements import check_achievements, get_achievements_with_status, get_achievement_progress
from analytics import (
    generate_pie_chart, generate_line_chart, generate_bar_chart,
    generate_monthly_trend_chart, get_smart_suggestions, get_detailed_stats
)

logger = logging.getLogger(__name__)

db: Database = None

def set_db(database):
    global db
    db = database

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    streaks = db.get_streaks(user_id)
    streak = streaks["current_streak"] if streaks else 0
    count = db.get_expense_count(user_id)
    xp = streaks["xp"] if streaks else 0
    level = streaks["level"] if streaks else 1
    streak_msg = f"\n🔥 Current Streak: {streak} days" if streak > 0 else ""
    msg = (
        f"👋 *Hi {update.effective_user.first_name}!*\n"
        f"💰 Expenses Logged: *{count}*\n"
        f"⭐ Level: *{level}* (XP: {xp}){streak_msg}\n\n"
        f"Send me an expense like:\n"
        f"`200 lunch` or `spent 500 on food`\n\n"
        f"Or tap a button below!"
    )
    if update.message:
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=main_reply_keyboard())
        await update.message.reply_text("🏠 *Main Menu*", parse_mode="Markdown", reply_markup=main_menu_keyboard())

# ----- Message Handler (Parse Expense) -----
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    user_id = update.effective_user.id

    parsed = parse_expense(text)
    if parsed:
        amount, category = parsed
        expense_id = db.add_expense(user_id, amount, category)
        comparison = fun_comparison(amount)
        emoji = get_category_emoji(category)
        new_ach = check_achievements(db, user_id)

        msg = (
            f"✅ *Expense Added!*\n\n"
            f"Amount: *{format_currency(amount)}*\n"
            f"Category: {emoji} {category}\n"
            f"Date: {format_date(datetime.now().strftime('%Y-%m-%d %H:%M:%S'))}"
        )
        if comparison:
            msg += f"\n\n💡 {comparison}"

        if new_ach:
            for key, label, desc in new_ach:
                msg += f"\n\n🏆 *New Achievement!* {label}\n_{desc}_"

        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=expense_confirmation_keyboard(expense_id))
    else:
        await update.message.reply_text(
            "🤔 *I couldn't understand that.*\n\n"
            "Try formats like:\n"
            "`200 lunch`\n"
            "`spent 500 on food`\n"
            "`Coffee 150`\n"
            "`Paid 450 for pizza`",
            parse_mode="Markdown"
        )

# ----- /summary -----
async def summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    data = db.get_monthly_summary(user_id)
    if not data:
        await update.message.reply_text("📭 No expenses this month.")
        return
    total = db.get_monthly_total(user_id)
    msg = f"📊 *This Month's Spending*\n\n"
    for row in data:
        pct = (row["total"] / total * 100) if total > 0 else 0
        bar = progress_bar(row["total"], total, 8)
        emoji = get_category_emoji(row["category"])
        msg += f"{emoji} {row['category']}: {format_currency(row['total'])}\n{bar} {pct:.0f}%\n\n"
    msg += f"──────────────\n*Total: {format_currency(total)}*"
    budgets = db.get_budgets(user_id)
    if budgets:
        msg += "\n\n📋 *Budgets:*\n"
        for b in budgets:
            spent = db.get_category_total(user_id, b["category"])
            pct = (spent / b["amount"] * 100) if b["amount"] > 0 else 0
            status = "✅" if spent <= b["amount"] else "⚠️"
            msg += f"{status} {b['category']}: {format_currency(spent)} / {format_currency(b['amount'])} ({pct:.0f}%)\n"
    await update.message.reply_text(msg, parse_mode="Markdown")

# ----- /total -----
async def total(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    total = db.get_monthly_total(user_id)
    await update.message.reply_text(f"💰 *Total spent this month: {format_currency(total)}*", parse_mode="Markdown")

# ----- /list -----
async def list_expenses(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    rows = db.get_all_expenses(user_id, 20)
    if not rows:
        await update.message.reply_text("📭 No expenses yet.")
        return
    msg = "📋 *Last 20 Expenses:*\n\n"
    for r in rows:
        emoji = get_category_emoji(r["category"])
        msg += f"`{r['id']:>3}` {emoji} {r['category']}: {format_currency(r['amount'])}\n    📅 {format_date(r['date'])}\n\n"
    await update.message.reply_text(msg, parse_mode="Markdown")

# ----- /delete -----
async def delete_expense(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Usage: /delete `<id>`\n\nFind the ID using /list", parse_mode="Markdown")
        return
    try:
        expense_id = int(context.args[0])
    except ValueError:
        await update.message.reply_text("❌ ID must be a number.")
        return
    user_id = update.effective_user.id
    if db.delete_expense(expense_id, user_id):
        await update.message.reply_text(f"🗑 Deleted expense #{expense_id}")
    else:
        await update.message.reply_text("❌ Expense not found.")

# ----- /from -----
async def date_range(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text
    match = re.search(r"from\s+(\d{4}-\d{2}-\d{2})\s+to\s+(\d{4}-\d{2}-\d{2})", text, re.IGNORECASE)
    if not match:
        await update.message.reply_text("Usage: /from 2025-01-01 to 2025-01-31", parse_mode="Markdown")
        return
    start, end = match.group(1), match.group(2)
    user_id = update.effective_user.id
    rows = db.get_expenses_by_date_range(user_id, start, end + " 23:59:59")
    if not rows:
        await update.message.reply_text(f"📭 No expenses from {start} to {end}.")
        return
    total = sum(r["amount"] for r in rows)
    msg = f"📅 *Expenses: {start} to {end}*\n\n"
    for r in rows:
        emoji = get_category_emoji(r["category"])
        msg += f"`{r['id']:>3}` {emoji} {r['category']}: {format_currency(r['amount'])}\n    📅 {format_date(r['date'])}\n\n"
    msg += f"──────────────\n*Total: {format_currency(total)}*"
    await update.message.reply_text(msg, parse_mode="Markdown")

# ----- /stats -----
async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    s = get_detailed_stats(db, user_id)
    if not s:
        await update.message.reply_text("📭 No expenses this month.")
        return
    streaks = db.get_streaks(user_id)
    streak = streaks["current_streak"] if streaks else 0
    msg = (
        f"📊 *Monthly Statistics*\n\n"
        f"💸 Total: *{format_currency(s['month_total'])}*\n"
        f"📝 Transactions: *{s['count']}*\n"
        f"📊 Average: *{format_currency(s['avg'])}*\n"
        f"🔺 Biggest: *{format_currency(s['biggest'])}*\n"
        f"📅 Today: *{format_currency(s['today_total'])}* ({s['today_count']} txns)\n"
        f"📆 This Week: *{format_currency(s['week_total'])}*\n"
        f"🔥 Streak: *{streak} days*"
    )
    await update.message.reply_text(msg, parse_mode="Markdown")

# ----- /categories -----
async def categories(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    cats = db.get_categories(user_id)
    if not cats:
        await update.message.reply_text("📭 No categories yet.")
        return
    msg = "📂 *All Categories:*\n\n"
    for c in cats:
        emoji = get_category_emoji(c)
        total = db.get_category_total(user_id, c)
        msg += f"{emoji} {c}: {format_currency(total)}\n"
    await update.message.reply_text(msg, parse_mode="Markdown")

# ----- /budget -----
async def budget(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not context.args:
        await update.message.reply_text("💰 *Budget Manager*", parse_mode="Markdown", reply_markup=budget_menu_keyboard())
        return
    if context.args[0] == "show":
        rows = db.get_budgets(user_id)
        if not rows:
            await update.message.reply_text("📭 No budgets set. Use `/budget set food 5000`", parse_mode="Markdown")
            return
        msg = "📋 *Your Budgets:*\n\n"
        for b in rows:
            spent = db.get_category_total(user_id, b["category"])
            pct = (spent / b["amount"] * 100) if b["amount"] > 0 else 0
            bar = progress_bar(spent, b["amount"])
            status = "✅" if spent <= b["amount"] else "⚠️"
            mood = "🟢" if pct <= 50 else ("🟡" if pct <= 80 else ("🟠" if pct <= 100 else "🔴"))
            msg += f"{status} {b['category']}: {format_currency(spent)} / {format_currency(b['amount'])}\n{mood} {bar} {pct:.0f}%\n\n"
        await update.message.reply_text(msg, parse_mode="Markdown")
    elif context.args[0] == "set" and len(context.args) >= 3:
        category = context.args[1].capitalize()
        try:
            amount = float(context.args[2])
        except ValueError:
            await update.message.reply_text("❌ Amount must be a number.")
            return
        db.set_budget(user_id, category, amount)
        await update.message.reply_text(f"✅ Budget set: {get_category_emoji(category)} {category} = {format_currency(amount)}", parse_mode="Markdown")
    elif context.args[0] == "delete" and len(context.args) >= 2:
        category = context.args[1].capitalize()
        db.delete_budget(user_id, category)
        await update.message.reply_text(f"🗑 Budget deleted for {category}.")
    else:
        await update.message.reply_text("Usage:\n/budget set food 5000\n/budget show\n/budget delete food", parse_mode="Markdown")

# ----- /export -----
async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("📤 *Export Options*", parse_mode="Markdown", reply_markup=export_keyboard())

async def export_csv(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    user_id = update.effective_user.id
    csv_data = db.export_csv(user_id)
    if not csv_data.strip() or csv_data == "id,amount,category,date,tags,note\n":
        msg_obj = update.message or update.callback_query.message
        await msg_obj.reply_text("📭 No expenses to export.")
        return
    f = io.StringIO(csv_data)
    msg_obj = update.message or update.callback_query.message
    await msg_obj.reply_document(document=f, filename="expenses.csv", caption="📄 Your expense data")

async def export_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    chat_id = update.effective_chat.id
    rows = db.cursor.execute("SELECT id, amount, category, date, tags, note FROM expenses WHERE user_id = ? ORDER BY date DESC", (user_id,)).fetchall()
    if not rows:
        msg_obj = update.message or update.callback_query.message
        await msg_obj.reply_text("📭 No expenses to export.")
        return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"
        headers = ["ID", "Amount", "Category", "Date", "Tags", "Note"]
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row_idx, r in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=r["id"])
            ws.cell(row=row_idx, column=2, value=r["amount"])
            ws.cell(row=row_idx, column=3, value=r["category"])
            ws.cell(row=row_idx, column=4, value=r["date"])
            ws.cell(row=row_idx, column=5, value=r["tags"])
            ws.cell(row=row_idx, column=6, value=r["note"])
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 20
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        msg_obj = update.message or update.callback_query.message
        await msg_obj.reply_document(document=buf, filename="expenses.xlsx", caption="📊 Excel report")
    except ImportError:
        msg_obj = update.message or update.callback_query.message
        await msg_obj.reply_text("❌ Excel export requires openpyxl. Use /export for CSV.")

async def export_pdf(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    rows = db.cursor.execute("SELECT amount, category, date FROM expenses WHERE user_id = ? ORDER BY date DESC LIMIT 50", (user_id,)).fetchall()
    if not rows:
        msg_obj = update.message or update.callback_query.message
        await msg_obj.reply_text("📭 No expenses to export.")
        return
    try:
        from fpdf import FPDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", "B", 16)
        pdf.cell(0, 10, "Expense Report", ln=True, align="C")
        pdf.ln(5)
        pdf.set_font("Helvetica", size=10)
        pdf.cell(10, 7, "#", border=1)
        pdf.cell(30, 7, "Amount", border=1)
        pdf.cell(40, 7, "Category", border=1)
        pdf.cell(50, 7, "Date", border=1)
        pdf.ln()
        total = 0
        for i, r in enumerate(rows, 1):
            pdf.cell(10, 7, str(i), border=1)
            pdf.cell(30, 7, f"Rs.{r['amount']:.2f}", border=1)
            pdf.cell(40, 7, r["category"], border=1)
            pdf.cell(50, 7, r["date"], border=1)
            pdf.ln()
            total += r["amount"]
        pdf.ln(5)
        pdf.set_font("Helvetica", "B", 12)
        pdf.cell(0, 10, f"Total: Rs.{total:.2f}", ln=True, align="R")
        buf = io.BytesIO()
        pdf.output(buf)
        buf.seek(0)
        msg_obj = update.message or update.callback_query.message
        await msg_obj.reply_document(document=buf, filename="expenses.pdf", caption="📑 PDF report")
    except ImportError:
        msg_obj = update.message or update.callback_query.message
        await msg_obj.reply_text("❌ PDF export requires fpdf2. Use /export for CSV.")

# ----- /reset -----
async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    db.delete_all_expenses(user_id)
    await update.message.reply_text("✨ *All your data has been wiped. Fresh start!*", parse_mode="Markdown")

# ----- /achievements -----
async def achievements(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    ach_list = get_achievements_with_status(db, user_id)
    progress = get_achievement_progress(db, user_id)
    streaks = db.get_streaks(user_id)
    xp = streaks["xp"] if streaks else 0
    level = streaks["level"] if streaks else 1
    msg = f"🏆 *Achievements*   ⭐ Level {level} (XP: {xp})\n\n"
    for key, label, desc, unlocked in ach_list:
        status = "✅" if unlocked else "❌"
        p = progress.get(key)
        if p and not unlocked:
            msg += f"{status} {label}\n   `{p[2]}`\n\n"
        else:
            msg += f"{status} {label}\n\n"
    await update.message.reply_text(msg, parse_mode="Markdown")

# ----- /fact -----
async def fun_fact(update: Update, context: ContextTypes.DEFAULT_TYPE):
    facts = [
        "If you save just ₹100 a month, that's ₹1200 in a year!",
        "People who track expenses save 15-30% more money.",
        "Small daily expenses add up fast - a ₹20 chai everyday is ₹600 a month!",
        "The 50/30/20 rule: 50% needs, 30% wants, 20% savings.",
        "Billionaires like Warren Buffett still live in the same house he bought in 1958.",
        "A study found people spend 44% more when using credit cards vs cash.",
        "The word 'salary' comes from 'salarium' - Roman soldiers were paid in salt!",
    ]
    import random
    await update.message.reply_text(f"💡 *Did you know?*\n{random.choice(facts)}", parse_mode="Markdown")

# ----- /search -----
async def search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Usage: /search coffee\n/search food\n/search >500", parse_mode="Markdown")
        return
    query = " ".join(context.args)
    user_id = update.effective_user.id
    if query.startswith(">"):
        try:
            min_amt = float(query[1:])
            rows = db.cursor.execute(
                "SELECT id, amount, category, date FROM expenses WHERE user_id = ? AND amount > ? ORDER BY date DESC LIMIT 20",
                (user_id, min_amt)
            ).fetchall()
        except ValueError:
            await update.message.reply_text("❌ Invalid number. Use: /search >500")
            return
    else:
        rows = db.search_expenses(user_id, query)
    if not rows:
        await update.message.reply_text(f"📭 No results for '{query}'.")
        return
    total = sum(r["amount"] for r in rows)
    msg = f"🔍 *Results for:* `{query}`\n\n"
    for r in rows:
        emoji = get_category_emoji(r["category"])
        msg += f"`{r['id']:>3}` {emoji} {r['category']}: {format_currency(r['amount'])}\n    📅 {format_date(r['date'])}\n\n"
    msg += f"*Total: {format_currency(total)}*"
    await update.message.reply_text(msg, parse_mode="Markdown")

# ----- /dashboard -----
async def dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    today_total = db.get_today_total(user_id)
    week_total = db.get_this_week_total(user_id)
    month_total = db.get_monthly_total(user_id)
    s = db.get_stats(user_id)
    count = s["count"] if s else 0
    biggest = s["biggest"] if s else 0
    streaks = db.get_streaks(user_id)
    streak = streaks["current_streak"] if streaks else 0
    xp = streaks["xp"] if streaks else 0
    level = streaks["level"] if streaks else 1
    budgets = db.get_budgets(user_id)
    budget_remaining = sum(b["amount"] for b in budgets) - sum(db.get_category_total(user_id, b["category"]) for b in budgets)
    ach_count = len(db.get_achievements(user_id))
    ach_total = 14
    ach_bar = progress_bar(ach_count, ach_total)
    msg = (
        f"📊 *Dashboard*\n\n"
        f"📅 Today: {format_currency(today_total)}\n"
        f"📆 This Week: {format_currency(week_total)}\n"
        f"📊 This Month: {format_currency(month_total)}\n\n"
        f"💰 Budget Remaining: {format_currency(max(0, budget_remaining))}\n"
        f"🔥 Streak: {streak} days\n"
        f"⭐ Level {level} (XP: {xp})\n"
        f"🏆 Achievements: {ach_bar} {ach_count}/{ach_total}\n\n"
        f"📝 Transactions: {count} | 🔺 Largest: {format_currency(biggest)}"
    )
    await update.message.reply_text(msg, parse_mode="Markdown")

# ----- /reminder -----
async def reminder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not context.args:
        r = db.get_reminder(user_id)
        status = "🔔 *On*" if r and r["enabled"] else "🔕 *Off*"
        time = r["reminder_time"] if r else "21:00"
        limit = r["daily_limit"] if r else 500
        msg = (
            f"⏰ *Daily Reminder*\n\n"
            f"Status: {status}\n"
            f"Time: {time}\n"
            f"Daily Limit: {format_currency(limit)}\n\n"
            f"Commands:\n"
            f"/reminder on\n"
            f"/reminder off\n"
            f"/reminder 21:30"
        )
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=reminder_keyboard())
        return
    arg = context.args[0].lower()
    if arg == "on":
        db.set_reminder(user_id, enabled=1)
        await update.message.reply_text("🔔 *Daily reminder enabled!*", parse_mode="Markdown")
    elif arg == "off":
        db.set_reminder(user_id, enabled=0)
        await update.message.reply_text("🔕 *Daily reminder disabled.*", parse_mode="Markdown")
    elif re.match(r"^\d{1,2}:\d{2}$", arg):
        db.set_reminder(user_id, reminder_time=arg)
        await update.message.reply_text(f"⏰ *Reminder time set to {arg}*", parse_mode="Markdown")
    else:
        await update.message.reply_text("Usage:\n/reminder on\n/reminder off\n/reminder 21:30")

# ----- /dailylimit -----
async def daily_limit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not context.args:
        await update.message.reply_text("Usage:\n/dailylimit 500\n/dailylimit 1000\n/dailylimit off", parse_mode="Markdown")
        return
    user_id = update.effective_user.id
    arg = context.args[0].lower()
    if arg == "off":
        db.set_daily_limit(user_id, 999999)
        await update.message.reply_text("✅ *Daily limit disabled.*", parse_mode="Markdown")
    else:
        try:
            limit = float(arg)
            db.set_daily_limit(user_id, limit)
            await update.message.reply_text(f"✅ *Daily limit set to {format_currency(limit)}*", parse_mode="Markdown")
        except ValueError:
            await update.message.reply_text("❌ Invalid amount.")

# ----- /help -----
async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = (
        "❓ *Help & Commands*\n\n"
        "*Expense:*\n"
        "`200 lunch` – Quick add\n"
        "`spent 500 on food` – Natural language\n"
        "`Coffee 150` – Auto-detect\n\n"
        "*Commands:*\n"
        "/summary – This month breakdown\n"
        "/total – Total this month\n"
        "/list – Last 20 expenses\n"
        "/delete `<id>` – Remove expense\n"
        "/from YYYY-MM-DD to YYYY-MM-DD – Filter\n"
        "/stats – Detailed statistics\n"
        "/categories – All categories\n"
        "/budget – Manage budgets\n"
        "/search `<query>` – Search expenses\n"
        "/dashboard – Full dashboard\n"
        "/export – CSV/Excel/PDF\n"
        "/reminder – Daily reminders\n"
        "/dailylimit `<amount>` – Set limit\n"
        "/goals – Savings goals\n"
        "/wishlist – Wishlist items\n"
        "/achievements – View badges\n"
        "/fact – Finance trivia\n"
        "/reset – Delete all data\n\n"
        "Tap the buttons below for interactive help!"
    )
    await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=help_keyboard())

# ----- /commands (list) -----
async def list_commands(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cmds = (
        "/start - Main menu\n"
        "/summary - Category breakdown this month\n"
        "/total - Total spent this month\n"
        "/list - Last 20 expenses\n"
        "/delete <id> - Delete an expense\n"
        "/from YYYY-MM-DD to YYYY-MM-DD - Filter by date\n"
        "/stats - Count, average, biggest expense\n"
        "/categories - All categories used\n"
        "/budget - Set or show budgets\n"
        "/search <query> - Search expenses\n"
        "/dashboard - Full spending dashboard\n"
        "/export - CSV, Excel, or PDF\n"
        "/reminder - Daily reminder settings\n"
        "/dailylimit <amount> - Set daily limit\n"
        "/goals - Savings goals\n"
        "/wishlist - Wishlist items\n"
        "/achievements - View your badges\n"
        "/fact - Random finance fact\n"
        "/help - Interactive help guide\n"
        "/reset - Delete all your data"
    )
    await update.message.reply_text(cmds)

# ----- /goals -----
async def goals(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not context.args:
        await update.message.reply_text("🎯 *Savings Goals*", parse_mode="Markdown", reply_markup=back_button("main_menu"))
        return
    if context.args[0] == "add" and len(context.args) >= 3:
        name = context.args[1]
        try:
            target = float(context.args[2])
        except ValueError:
            await update.message.reply_text("❌ Target must be a number.")
            return
        db.add_goal(user_id, name, target)
        await update.message.reply_text(f"🎯 Goal set: {name} = {format_currency(target)}")
    elif context.args[0] == "show":
        goals_list = db.get_goals(user_id)
        if not goals_list:
            await update.message.reply_text("📭 No goals. Use /goals add MacBook 120000")
            return
        msg = "🎯 *Your Goals*\n\n"
        for g in goals_list:
            pct = (g["saved"] / g["target"] * 100) if g["target"] > 0 else 0
            bar = progress_bar(g["saved"], g["target"])
            msg += f"*{g['name']}*\n"
            msg += f"Target: {format_currency(g['target'])}\n"
            msg += f"Saved: {format_currency(g['saved'])}\n"
            msg += f"{bar} {pct:.0f}%\n\n"
        await update.message.reply_text(msg, parse_mode="Markdown")
    else:
        await update.message.reply_text("Usage:\n/goals add MacBook 120000\n/goals show")

# ----- /wishlist -----
async def wishlist(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if not context.args:
        await update.message.reply_text("🎁 *Wishlist*", parse_mode="Markdown", reply_markup=back_button("main_menu"))
        return
    if context.args[0] == "add" and len(context.args) >= 3:
        name = context.args[1]
        try:
            price = float(context.args[2])
        except ValueError:
            await update.message.reply_text("❌ Price must be a number.")
            return
        db.add_wishlist_item(user_id, name, price)
        await update.message.reply_text(f"🎁 Added to wishlist: {name} = {format_currency(price)}")
    elif context.args[0] == "show":
        items = db.get_wishlist(user_id)
        if not items:
            await update.message.reply_text("📭 No wishlist items.")
            return
        msg = "🎁 *Wishlist*\n\n"
        for item in items:
            pct = (item["saved"] / item["price"] * 100) if item["price"] > 0 else 0
            bar = progress_bar(item["saved"], item["price"])
            months_needed = max(1, int((item["price"] - item["saved"]) / 1000)) if (item["price"] - item["saved"]) > 0 else 0
            msg += f"*{item['name']}*\n"
            msg += f"Price: {format_currency(item['price'])}\n"
            msg += f"Saved: {format_currency(item['saved'])}\n"
            msg += f"{bar} {pct:.0f}%\n"
            msg += f"⏱ ~{months_needed} months at ₹1000/month\n\n"
        await update.message.reply_text(msg, parse_mode="Markdown")
    else:
        await update.message.reply_text("Usage:\n/wishlist add Laptop 70000\n/wishlist show")

# ----- Callback Query Handler -----
async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    data = query.data
    user_id = query.from_user.id

    if data == "main_menu":
        streaks = db.get_streaks(user_id)
        streak = streaks["current_streak"] if streaks else 0
        count = db.get_expense_count(user_id)
        msg = f"🏠 *Main Menu*\n📝 {count} expenses logged" + (f"\n🔥 Streak: {streak}d" if streak > 0 else "")
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=main_menu_keyboard())

    elif data == "add_expense":
        await query.edit_message_text(
            "➕ *Add Expense*\n\nSend an expense in any format:\n"
            "`200 lunch`\n`spent 500 on food`\n`Coffee 150`\n`Paid 450 for pizza`",
            parse_mode="Markdown", reply_markup=back_button("main_menu")
        )

    elif data == "analytics_menu":
        await query.edit_message_text("📊 *Analytics*", parse_mode="Markdown", reply_markup=analytics_keyboard())

    elif data in ("analytics_today", "analytics_week", "analytics_month", "analytics_lastmonth"):
        period_map = {"analytics_today": "today", "analytics_week": "week", "analytics_month": "month", "analytics_lastmonth": "lastmonth"}
        period = period_map[data]
        start, end = get_date_range(period)
        rows = db.get_expenses_by_date_range(user_id, start, end + " 23:59:59")
        if not rows:
            await query.edit_message_text(f"📭 No expenses for this period.", reply_markup=back_button("analytics_menu"))
            return
        total = sum(r["amount"] for r in rows)
        count = len(rows)
        avg = total / count if count > 0 else 0
        biggest = max(r["amount"] for r in rows)
        smallest = min(r["amount"] for r in rows)
        cat_totals = {}
        for r in rows:
            cat_totals[r["category"]] = cat_totals.get(r["category"], 0) + r["amount"]
        top_cat = max(cat_totals, key=cat_totals.get) if cat_totals else "N/A"
        period_name = period.capitalize()
        msg = (
            f"📊 *{period_name} Summary*\n\n"
            f"💸 Total: *{format_currency(total)}*\n"
            f"📝 Transactions: *{count}*\n"
            f"📊 Average: *{format_currency(avg)}*\n"
            f"🔺 Max: *{format_currency(biggest)}*\n"
            f"🔻 Min: *{format_currency(smallest)}*\n"
            f"🏆 Top Category: {get_category_emoji(top_cat)} {top_cat}\n\n"
            f"*Category Breakdown:*\n"
        )
        for cat, amt in sorted(cat_totals.items(), key=lambda x: x[1], reverse=True):
            pct = (amt / total * 100) if total > 0 else 0
            bar = progress_bar(amt, total, 6)
            msg += f"{get_category_emoji(cat)} {cat}: {format_currency(amt)}\n{bar} {pct:.0f}%\n\n"
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=back_button("analytics_menu"))

    elif data == "analytics_custom":
        await query.edit_message_text(
            "📅 Enter a date range:\n/from 2025-01-01 to 2025-01-31",
            parse_mode="Markdown", reply_markup=back_button("analytics_menu")
        )

    elif data == "reports_menu":
        await query.edit_message_text("📅 *Reports & Charts*", parse_mode="Markdown", reply_markup=reports_keyboard())

    elif data in ("chart_pie", "chart_line", "chart_bar", "chart_trend"):
        await query.edit_message_text("📊 *Generating chart...*", parse_mode="Markdown")
        chart_map = {
            "chart_pie": ("📊 Pie Chart", generate_pie_chart),
            "chart_line": ("📈 Line Chart", generate_line_chart),
            "chart_bar": ("📊 Bar Chart", generate_bar_chart),
            "chart_trend": ("📅 Monthly Trend", generate_monthly_trend_chart),
        }
        title, gen_func = chart_map[data]
        buf = gen_func(db, user_id)
        if buf:
            await query.delete_message()
            await context.bot.send_photo(chat_id=user_id, photo=buf, caption=title)
            streaks = db.get_streaks(user_id)
            streak = streaks["current_streak"] if streaks else 0
            count = db.get_expense_count(user_id)
            menu_msg = f"🏠 *Main Menu*\n📝 {count} expenses logged" + (f"\n🔥 Streak: {streak}d" if streak > 0 else "")
            await context.bot.send_message(chat_id=user_id, text=menu_msg, parse_mode="Markdown", reply_markup=main_menu_keyboard())
        else:
            msg = (f"{title}\n\n"
                   "❌ Could not generate chart.\n"
                   "📭 No data for this month or matplotlib unavailable.")
            await query.edit_message_text(msg, reply_markup=back_button("reports_menu"))

    elif data == "budget_menu":
        await query.edit_message_text("💰 *Budget Manager*", parse_mode="Markdown", reply_markup=budget_menu_keyboard())

    elif data == "budget_show":
        rows = db.get_budgets(user_id)
        if not rows:
            await query.edit_message_text("📭 No budgets set.", reply_markup=back_button("budget_menu"))
            return
        msg = "📋 *Your Budgets:*\n\n"
        for b in rows:
            spent = db.get_category_total(user_id, b["category"])
            pct = (spent / b["amount"] * 100) if b["amount"] > 0 else 0
            bar = progress_bar(spent, b["amount"])
            mood = "🟢" if pct <= 50 else ("🟡" if pct <= 80 else ("🟠" if pct <= 100 else "🔴"))
            msg += f"{mood} {b['category']}: {format_currency(spent)} / {format_currency(b['amount'])}\n{bar} {pct:.0f}%\n\n"
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=back_button("budget_menu"))

    elif data == "budget_set":
        await query.edit_message_text(
            "✏️ Set a budget:\n`/budget set food 5000`\n`/budget set travel 3000`",
            parse_mode="Markdown", reply_markup=back_button("budget_menu")
        )

    elif data == "budget_delete":
        await query.edit_message_text(
            "🗑 Delete a budget:\n`/budget delete food`",
            parse_mode="Markdown", reply_markup=back_button("budget_menu")
        )

    elif data == "settings_menu":
        await query.edit_message_text("⚙️ *Settings*", parse_mode="Markdown", reply_markup=settings_keyboard())

    elif data == "setting_dailylimit":
        current = db.get_reminder(user_id)
        limit = current["daily_limit"] if current else 500
        await query.edit_message_text(
            f"💰 Current daily limit: {format_currency(limit)}\n\n"
            f"Change it:\n`/dailylimit 500`\n`/dailylimit 1000`\n`/dailylimit off`",
            parse_mode="Markdown", reply_markup=back_button("settings_menu")
        )

    elif data == "setting_remindertime":
        current = db.get_reminder(user_id)
        t = current["reminder_time"] if current else "21:00"
        await query.edit_message_text(
            f"⏰ Current reminder time: {t}\n\n"
            f"Change it:\n`/reminder 21:30`\n`/reminder 20:00`",
            parse_mode="Markdown", reply_markup=back_button("settings_menu")
        )

    elif data == "setting_currency":
        await query.edit_message_text(
            "🔄 Currency setting coming soon!\nCurrently using ₹ (INR).",
            reply_markup=back_button("settings_menu")
        )

    elif data == "setting_darkmode":
        current = db.get_setting(user_id, "dark_mode")
        new_val = 0 if current else 1
        db.set_setting(user_id, "dark_mode", new_val)
        status = "🌙 *Enabled*" if new_val else "☀️ *Disabled*"
        await query.edit_message_text(f"Dark mode: {status}", parse_mode="Markdown", reply_markup=back_button("settings_menu"))

    elif data == "help_menu":
        await query.edit_message_text("❓ *Interactive Help*", parse_mode="Markdown", reply_markup=help_keyboard())

    elif data in ("help_expense", "help_budget", "help_reports", "help_settings", "help_achievements"):
        help_texts = {
            "help_expense": (
                "💸 *Expense Tips*\n\n"
                "• `200 lunch` – Quick add\n"
                "• `spent 500 on food` – Full sentence\n"
                "• `Coffee 150` – Auto-detects coffee category\n"
                "• `Paid 450 for pizza` – Also works\n\n"
                "After adding, you can Undo or Edit the expense."
            ),
            "help_budget": (
                "💰 *Budget Help*\n\n"
                "• `/budget set food 5000` – Set budget\n"
                "• `/budget show` – View all budgets\n"
                "• `/budget delete food` – Remove budget\n\n"
                "You'll get alerts when you cross 80%, 100%, and 120%."
            ),
            "help_reports": (
                "📊 *Reports Help*\n\n"
                "• Pie Chart – Category breakdown\n"
                "• Line Chart – Daily trend\n"
                "• Bar Chart – Category comparison\n"
                "• Monthly trend – Track spending over time\n\n"
                "You also get weekly (Sun) and monthly (1st) auto-reports."
            ),
            "help_settings": (
                "⚙️ *Settings Help*\n\n"
                "• /dailylimit 500 – Set your daily budget\n"
                "• /reminder on/off – Toggle 9 PM reminder\n"
                "• /reminder 21:30 – Custom reminder time\n\n"
                "The bot alerts you if you exceed your daily limit!"
            ),
            "help_achievements": (
                "🏆 *Achievements Help*\n\n"
                "Earn badges by:\n"
                "• Logging expenses (1, 5, 10, 25, 50, 100)\n"
                "• Maintaining streaks (3, 7, 14, 30 days)\n"
                "• Setting budgets and goals\n"
                "• Logging on weekends or early mornings\n\n"
                "Each achievement gives you +50 XP!"
            ),
        }
        await query.edit_message_text(help_texts[data], parse_mode="Markdown", reply_markup=back_button("help_menu"))

    elif data == "show_achievements":
        ach_list = get_achievements_with_status(db, user_id)
        progress = get_achievement_progress(db, user_id)
        streaks = db.get_streaks(user_id)
        xp = streaks["xp"] if streaks else 0
        level = streaks["level"] if streaks else 1
        msg = f"🏆 *Achievements*   ⭐ Level {level} (XP: {xp})\n\n"
        for key, label, desc, unlocked in ach_list:
            status = "✅" if unlocked else "❌"
            p = progress.get(key)
            if p and not unlocked:
                msg += f"{status} {label}\n   `{p[2]}`\n\n"
            else:
                msg += f"{status} {label}\n\n"
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=back_button("main_menu"))

    elif data == "goal_add":
        await query.edit_message_text(
            "🎯 Add a savings goal:\n`/goals add MacBook 120000`",
            parse_mode="Markdown", reply_markup=back_button("main_menu")
        )

    elif data == "goal_view":
        goals_list = db.get_goals(user_id)
        if not goals_list:
            await query.edit_message_text("📭 No goals yet.", reply_markup=back_button("main_menu"))
            return
        msg = "🎯 *Your Goals*\n\n"
        for g in goals_list:
            pct = (g["saved"] / g["target"] * 100) if g["target"] > 0 else 0
            bar = progress_bar(g["saved"], g["target"])
            msg += f"*{g['name']}*\nTarget: {format_currency(g['target'])}\nSaved: {format_currency(g['saved'])}\n{bar} {pct:.0f}%\n\n"
        await query.edit_message_text(msg, parse_mode="Markdown", reply_markup=back_button("main_menu"))

    elif data in ("reminder_on", "reminder_off"):
        enabled = 1 if data == "reminder_on" else 0
        db.set_reminder(user_id, enabled=enabled)
        status = "🔔 *Enabled*" if enabled else "🔕 *Disabled*"
        await query.edit_message_text(f"Daily reminder: {status}", parse_mode="Markdown", reply_markup=reminder_keyboard())

    elif data == "reminder_time":
        await query.edit_message_text(
            "⏰ Set reminder time:\n`/reminder 20:00`\n`/reminder 21:30`",
            parse_mode="Markdown", reply_markup=reminder_keyboard()
        )

    elif data.startswith("undo_"):
        expense_id = int(data.split("_")[1])
        if db.delete_expense(expense_id, user_id):
            await query.edit_message_text(f"↩️ *Expense #{expense_id} undone!*", parse_mode="Markdown", reply_markup=back_button("main_menu"))
        else:
            await query.edit_message_text("❌ Could not undo.", reply_markup=back_button("main_menu"))

    elif data.startswith("edit_"):
        expense_id = int(data.split("_")[1])
        expense = db.get_expense(expense_id, user_id)
        if expense:
            await query.edit_message_text(
                f"✏️ *Editing expense #{expense_id}*\n\n"
                f"Current: {format_currency(expense['amount'])} on {expense['category']}\n\n"
                f"To edit, send:\n`/edit {expense_id} 300` – change amount\n`/edit {expense_id} 300 food` – change both",
                parse_mode="Markdown", reply_markup=back_button("main_menu")
            )
        else:
            await query.edit_message_text("❌ Expense not found.", reply_markup=back_button("main_menu"))

    elif data in ("export_csv", "export_excel", "export_pdf"):
        await query.edit_message_text("📤 *Generating export...*", parse_mode="Markdown")
        if data == "export_csv":
            await export_csv(update, context)
        elif data == "export_excel":
            await export_excel(update, context)
        elif data == "export_pdf":
            await export_pdf(update, context)
        # Send main menu back
        streaks = db.get_streaks(user_id)
        streak = streaks["current_streak"] if streaks else 0
        count = db.get_expense_count(user_id)
        menu_msg = f"🏠 *Main Menu*\n📝 {count} expenses logged" + (f"\n🔥 Streak: {streak}d" if streak > 0 else "")
        await context.bot.send_message(chat_id=user_id, text=menu_msg, parse_mode="Markdown", reply_markup=main_menu_keyboard())

    else:
        await query.edit_message_text(f"❓ Unknown action: {data}", reply_markup=back_button("main_menu"))

async def edit_expense(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if len(context.args) < 2:
        await update.message.reply_text("Usage: /edit <id> <amount> [category]", parse_mode="Markdown")
        return
    try:
        expense_id = int(context.args[0])
        amount = float(context.args[1])
    except ValueError:
        await update.message.reply_text("❌ Invalid ID or amount.")
        return
    category = context.args[2].capitalize() if len(context.args) >= 3 else None
    user_id = update.effective_user.id
    if db.update_expense(expense_id, user_id, amount, category):
        await update.message.reply_text(f"✏️ *Expense #{expense_id} updated!*", parse_mode="Markdown")
    else:
        await update.message.reply_text("❌ Expense not found.")
